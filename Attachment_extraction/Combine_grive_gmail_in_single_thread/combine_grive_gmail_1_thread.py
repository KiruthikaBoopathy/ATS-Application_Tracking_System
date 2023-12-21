# import concurrent
# import time
# import email
# import imaplib
# from concurrent import futures
# from email.header import decode_header
# from email.utils import parsedate_to_datetime
# import html2text
# import json
# import fitz  # PyMuPDF
# import io
# import openpyxl
# from docx import Document
#
# class Email_extract():
#     def get_emails(self,imapUserEmail, imapPassword):
#         imap_server = "imap.gmail.com"
#         imap_port = 993
#
#         imap_conn = imaplib.IMAP4_SSL(imap_server, imap_port)
#
#         imap_conn.login(imapUserEmail, imapPassword)
#
#         imap_conn.select('inbox')
#
#         _, message_ids = imap_conn.search(None, "UNSEEN")
#
#         emails = []
#
#         for message_id in message_ids[0].split():
#             try:
#                 _, data = imap_conn.fetch(message_id, "(RFC822)")
#                 raw_email = data[0][1]
#                 email_message = email.message_from_bytes(raw_email)
#
#                 subject = email_message["Subject"]
#                 sender = email.utils.parseaddr(email_message["From"])[0]
#                 body1 = ""
#                 attachments = []
#
#                 date_received = parsedate_to_datetime(email_message["Date"])
#
#                 if email_message.is_multipart():
#                     for part in email_message.walk():
#                         if part.get_content_type() == "text/plain":
#                             body1 = part.get_payload(decode=True).decode()
#                         elif part.get_content_type() == "text/html":
#                             body1 = html2text.html2text(part.get_payload(decode=True).decode())
#                         elif part.get_content_maintype() != 'multipart' and part.get('Content-Disposition') is not None:
#                             filename = part.get_filename()
#                             if filename:
#                                 filename, encoding = decode_header(filename)[0]
#                                 if isinstance(filename, bytes):
#                                     filename = filename.decode(encoding or "utf-8")
#                                 attachment_data = part.get_payload(decode=True)
#                                 attachments.append({
#                                     "filename": filename,
#                                     "data": attachment_data
#                                 })
#
#                 email_data = {
#                     "subject": subject,
#                     "sender": sender,
#                     "body": body1,
#                     "date_received": date_received,
#                     "email_id": message_id,
#                     "attachments": attachments
#                 }
#
#                 emails.append(email_data)
#             except Exception as e:
#                 print(f"Error processing email: {str(e)}")
#
#                 imap_conn.store(message_id, '-FLAGS', '(\Seen)')
#
#         imap_conn.close()
#
#         return emails
#
#
#
#     def op_process_mail(self,email_data):
#         # get_emails = self.get_emails(imapUserEmail, imapPassword)
#         subject = email_data["subject"]
#         sender = email_data["sender"]
#         date_received = email_data["date_received"]
#         attachments = email_data["attachments"]
#         body = email_data["body"]
#
#         print("Subject:", subject)
#         print("Sender:", sender)
#         print("Date Received:", date_received)
#
#         if attachments:
#             for attachment in attachments:
#                 filename = attachment["filename"]
#                 file_format = filename.split('.')[-1].lower()
#                 content = attachment["data"]
#
#                 if file_format == 'xlsx':
#                     wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
#                     text_content = ""
#                     for sheet_name in wb.sheetnames:
#                         sheet = wb[sheet_name]
#                         for row in sheet.iter_rows():
#                             row_values = [str(cell.value) for cell in row if cell.value is not None]
#                             text_content += " ".join(row_values) + "\n"
#                         text_content += "\n"
#
#                     print("Attachments:")
#                     print(json.dumps({
#                         "filename": filename,
#                         "content_type": "text_from_xlsx",
#                         "content": text_content
#                     }, indent=2))
#
#                 elif file_format == 'pdf':
#                     pdf_document = fitz.open(stream=io.BytesIO(content), filetype="pdf")
#                     text_content = ""
#                     for page_number in range(pdf_document.page_count):
#                         page = pdf_document[page_number]
#                         text_content += page.get_text()
#
#                     print("Attachments:")
#                     print(json.dumps({
#                         "filename": filename,
#                         "content_type": "text_from_pdf",
#                         "content": text_content
#                     }, indent=2))
#
#                 elif file_format == 'docx':
#                     doc = Document(io.BytesIO(content))
#                     text_content = "\n".join([paragraph.text for paragraph in doc.paragraphs])
#
#                     print("Attachments:")
#                     print(json.dumps({
#                         "filename": filename,
#                         "content_type": "text_from_docx",
#                         "content": text_content
#                     }, indent=2))
#
#                 elif file_format == 'py':
#                     text_content = content.decode('utf-8')
#
#                     print("Attachments:")
#                     print(json.dumps({
#                         "filename": filename,
#                         "content_type": "text_from_python",
#                         "content": text_content
#                     }, indent=2))
#
#         else:
#             if body:
#                 print("Body:", body)
#             else:
#                 print("No Body content")
#
#         print("\n" + "*" * 100 + "\n")
#
#


import os
from concurrent import futures

import fitz  # PyMuPDF
import io
import tempfile
import time
import openpyxl
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.http import MediaIoBaseDownload
import concurrent.futures


class Gdrive_Extract():
# Set the API key file path
    API_KEY_FILE = r"C:\Users\Vrdella\Downloads\gdrive_credentials.json"

    # Set the OAuth scope and redirect URI
    SCOPES = ['https://www.googleapis.com/auth/drive.metadata.readonly']

    # Create credentials using the API key file and OAuth
    credentials = None
    if os.path.exists('token.json'):
        credentials = Credentials.from_authorized_user_file('token.json')
    if not credentials or not credentials.valid:
        if credentials and credentials.expired and credentials.refresh_token:
            credentials.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(API_KEY_FILE, SCOPES)
            credentials = flow.run_local_server(port=0)

        with open('token.json', 'w') as token:
            token.write(credentials.to_json())

    # Build the Google Drive API service
    drive_service = build('drive', 'v3', credentials=credentials)

    # List folders in the root directory, sorted by name
    results = drive_service.files().list(
        q="mimeType='application/vnd.google-apps.folder' and trashed=false",
        pageSize=10, fields="nextPageToken, files(id, name)",
        orderBy="name").execute()
    folders = results.get('files', [])


    def content_extract(self,folders,drive_service):
        if not folders:
            print('No folders found.')
        else:
            # print('Folders:')
            for folder in folders:
                folder_name = folder["name"]
                folder_id = folder["id"]
                # print(f'Processing folder: {folder_name} ({folder_id})')

                # List files inside the folder, sorted by name
                files_results = drive_service.files().list(
                    q=f"'{folder_id}' in parents and trashed=false",
                    pageSize=10, fields="nextPageToken, files(id, name, mimeType)",
                    orderBy="name").execute()
                files = files_results.get('files', [])

                # def content_extraction(files):
                if not files:
                    print(f'No files found inside the folder "{folder_name}".')
                else:

                    # def content_extraction(files):
                    print('\n')
                    print(f'Files inside the folder "{folder_name}":')
                    for file in files:
                        file_name = file["name"]
                        file_id = file["id"]
                        mime_type = file['mimeType']

                        print(f' {file_name} ({file_id})')

                        # print(f'Processing file: {file_name} (ID: {file_id})')
                        request = drive_service.files().get_media(fileId=file_id)
                        file_path = os.path.join(os.getcwd(), file_name)
                        # print(file_path)

                        # Download the file
                        fh = io.FileIO(file_path, 'wb')
                        downloader = MediaIoBaseDownload(fh, request)
                        done = False
                        while not done:
                            status, done = downloader.next_chunk()

                        # Extract text content based on MIME type
                        if file_name.lower().endswith('.xlsx'):
                            workbook = openpyxl.load_workbook(file_path)

                            excel_text = ""

                            # Iterate through all sheets in the workbook
                            for sheet_name in workbook.sheetnames:
                                sheet = workbook[sheet_name]
                                excel_text += f"Sheet: {sheet_name}\n"

                                # Iterate through all rows in the sheet
                                for row in sheet.iter_rows(values_only=True):
                                    row_text = "\t".join(str(cell) for cell in row)
                                    excel_text += row_text + "\n"

                            print(f'Excel Content:\n{excel_text}\n{"=" * 50}\n')
                            fh.close()
                            os.remove(file_path)

                        elif file_name.lower().endswith('.pdf'):
                            pdf_document = fitz.open(file_path)
                            pdf_text = ""

                            for page_number in range(pdf_document.page_count):
                                page = pdf_document[page_number]
                                pdf_text += page.get_text()

                            pdf_document.close()
                            print(f'PDF Content:\n{pdf_text}\n{"=" * 50}\n')
                            fh.close()
                            os.remove(file_path)

                        elif file_name.lower().endswith('.py'):
                            # Assuming it's a Python file, read its content
                            with open(file_path, 'r', encoding='utf-8') as python_file:
                                python_code = python_file.read()

                            print(f'Python File Content:\n{python_code}\n{"=" * 50}\n')
                            # Close the file
                            fh.close()
                            os.remove(file_path)


if __name__ == "main":
    # start = time.time()

    # imapUserEmail = "kiruthika.b@vrdella.com"
    # imapPassword = "renm kixf nlxy avbx"
    #
    # instance = Email_extract()
    # emails = instance.get_emails(imapUserEmail, imapPassword)
    # for email_data in emails:
    #         instance.op_process_mail(email_data)
    #
    # end = time.time()
    # print("Total Execution Time:", end - start)


    s=time.time()
    instance1=Gdrive_Extract()
    instance1.content_extract(instance1.folders, instance1.drive_service)
    e=time.time()

    if __name__ == "__main__":
        gdrive_instance = Gdrive_Extract()

        # Fetch folders here
        results = gdrive_instance.drive_service.files().list(
            q="mimeType='application/vnd.google-apps.folder' and trashed=false",
            pageSize=10, fields="nextPageToken, files(id, name)",
            orderBy="name").execute()
        folders = results.get('files', [])

        s = time.time()
        gdrive_instance.content_extract(folders)
        e = time.time()
        print(e - s)

        







