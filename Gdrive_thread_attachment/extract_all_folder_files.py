# import os
# from concurrent import futures
# 
# import fitz  # PyMuPDF
# import io
# import tempfile
# import time
# import openpyxl
# from googleapiclient.discovery import build
# from google.oauth2.credentials import Credentials
# from google_auth_oauthlib.flow import InstalledAppFlow
# from google.auth.transport.requests import Request
# from googleapiclient.http import MediaIoBaseDownload
# import concurrent.futures
# 
# # Set the API key file path
# API_KEY_FILE = r"C:\Users\Vrdella\Downloads\gdrive_credentials.json"
# 
# # Set the OAuth scope and redirect URI
# SCOPES = ['https://www.googleapis.com/auth/drive.metadata.readonly']
# 
# # Create credentials using the API key file and OAuth
# credentials = None
# if os.path.exists('token.json'):
#     credentials = Credentials.from_authorized_user_file('token.json')
# if not credentials or not credentials.valid:
#     if credentials and credentials.expired and credentials.refresh_token:
#         credentials.refresh(Request())
#     else:
#         flow = InstalledAppFlow.from_client_secrets_file(API_KEY_FILE, SCOPES)
#         credentials = flow.run_local_server(port=0)
# 
#     with open('token.json', 'w') as token:
#         token.write(credentials.to_json())
# 
# # Build the Google Drive API service
# drive_service = build('drive', 'v3', credentials=credentials)
# 
# # List folders in the root directory, sorted by name
# results = drive_service.files().list(
#     q="mimeType='application/vnd.google-apps.folder' and trashed=false",
#     pageSize=10, fields="nextPageToken, files(id, name)",
#     orderBy="name").execute()
# folders = results.get('files', [])
# 
# 
# def content_extract(folders):
#     if not folders:
#         print('No folders found.')
#     else:
#         # print('Folders:')
#         for folder in folders:
#             folder_name = folder["name"]
#             folder_id = folder["id"]
#             # print(f'Processing folder: {folder_name} ({folder_id})')
# 
#             # List files inside the folder, sorted by name
#             files_results = drive_service.files().list(
#                 q=f"'{folder_id}' in parents and trashed=false",
#                 pageSize=10, fields="nextPageToken, files(id, name, mimeType)",
#                 orderBy="name").execute()
#             files = files_results.get('files', [])
# 
#             # def content_extraction(files):
#             if not files:
#                 print(f'No files found inside the folder "{folder_name}".')
#             else:
# 
#                 # def content_extraction(files):
#                 print('\n')
#                 print(f'Files inside the folder "{folder_name}":')
#                 for file in files:
#                     file_name = file["name"]
#                     file_id = file["id"]
#                     mime_type = file['mimeType']
# 
#                     print(f' {file_name} ({file_id})')
# 
#                     # print(f'Processing file: {file_name} (ID: {file_id})')
#                     request = drive_service.files().get_media(fileId=file_id)
#                     file_path = os.path.join(os.getcwd(), file_name)
#                     # print(file_path)
# 
#                     # Download the file
#                     fh = io.FileIO(file_path, 'wb')
#                     downloader = MediaIoBaseDownload(fh, request)
#                     done = False
#                     while not done:
#                         status, done = downloader.next_chunk()
# 
#                     # Extract text content based on MIME type
#                     if file_name.lower().endswith('.xlsx'):
#                         workbook = openpyxl.load_workbook(file_path)
# 
#                         excel_text = ""
# 
#                         # Iterate through all sheets in the workbook
#                         for sheet_name in workbook.sheetnames:
#                             sheet = workbook[sheet_name]
#                             excel_text += f"Sheet: {sheet_name}\n"
# 
#                             # Iterate through all rows in the sheet
#                             for row in sheet.iter_rows(values_only=True):
#                                 row_text = "\t".join(str(cell) for cell in row)
#                                 excel_text += row_text + "\n"
# 
#                         print(f'Excel Content:\n{excel_text}\n{"=" * 50}\n')
#                         fh.close()
#                         os.remove(file_path)
# 
#                     elif file_name.lower().endswith('.pdf'):
#                         pdf_document = fitz.open(file_path)
#                         pdf_text = ""
# 
#                         for page_number in range(pdf_document.page_count):
#                             page = pdf_document[page_number]
#                             pdf_text += page.get_text()
# 
#                         pdf_document.close()
#                         print(f'PDF Content:\n{pdf_text}\n{"=" * 50}\n')
#                         fh.close()
#                         os.remove(file_path)
# 
#                     elif file_name.lower().endswith('.py'):
#                         # Assuming it's a Python file, read its content
#                         with open(file_path, 'r', encoding='utf-8') as python_file:
#                             python_code = python_file.read()
# 
#                         print(f'Python File Content:\n{python_code}\n{"=" * 50}\n')
#                         # Close the file
#                         fh.close()
#                         os.remove(file_path)
# 
# 
# def run_parallel(folders):
#     with concurrent.futures.ThreadPoolExecutor() as executor:
#         executor.map(content_extract, folders)
# 
# s=time.time()
# run_parallel(folders)
# content_extract(folders)
# e=time.time()
# print(e-s)
# 


import os
from concurrent import futures
import fitz
import io
import openpyxl
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.http import MediaIoBaseDownload
import concurrent.futures
import time


class GoogleDriveExtractor:
    def __init__(self):
        # Set the API key file path
        self.API_KEY_FILE = r"C:\Users\Vrdella\Downloads\gdrive_credentials.json"

        # Set the OAuth scope and redirect URI
        self.SCOPES = ['https://www.googleapis.com/auth/drive.metadata.readonly']

        # Create credentials using the API key file and OAuth
        self.credentials = None
        if os.path.exists('token.json'):
            self.credentials = Credentials.from_authorized_user_file('token.json')
        if not self.credentials or not self.credentials.valid:
            if self.credentials and self.credentials.expired and self.credentials.refresh_token:
                self.credentials.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(self.API_KEY_FILE, self.SCOPES)
                self.credentials = flow.run_local_server(port=0)

            with open('token.json', 'w') as token:
                token.write(self.credentials.to_json())

        # Build the Google Drive API service
        self.drive_service = build('drive', 'v3', credentials=self.credentials)

    def list_folders(self):
        # List folders in the root directory, sorted by name
        results = self.drive_service.files().list(
            q="mimeType='application/vnd.google-apps.folder' and trashed=false",
            pageSize=10, fields="nextPageToken, files(id, name)",
            orderBy="name").execute()
        folders = results.get('files', [])
        return folders

    def content_extract(self, folders):
        if not folders:
            print('No folders found.')
        else:
            for folder in folders:
                folder_name = folder["name"]
                folder_id = folder["id"]

                # List files inside the folder, sorted by name
                files_results = self.drive_service.files().list(
                    q=f"'{folder_id}' in parents and trashed=false",
                    pageSize=10, fields="nextPageToken, files(id, name, mimeType)",
                    orderBy="name").execute()
                files = files_results.get('files', [])

                if not files:
                    print(f'No files found inside the folder "{folder_name}".')
                else:
                    print(f'\nFiles inside the folder "{folder_name}":')
                    for file in files:
                        file_name = file["name"]
                        file_id = file["id"]
                        mime_type = file['mimeType']

                        print(f' {file_name} ({file_id})')

                        request = self.drive_service.files().get_media(fileId=file_id)
                        file_path = os.path.join(os.getcwd(), file_name)

                        fh = io.FileIO(file_path, 'wb')
                        downloader = MediaIoBaseDownload(fh, request)
                        done = False
                        while not done:
                            status, done = downloader.next_chunk()

                        self.extract_content(file_name, file_path)

    def extract_content(self, file_name, file_path, fh):
        if file_name.lower().endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_path)
            excel_text = ""

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                excel_text += f"Sheet: {sheet_name}\n"

                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) for cell in row)
                    excel_text += row_text + "\n"

            print(f'Excel Content:\n{excel_text}\n{"=" * 50}\n')
            fh.close()
            os.remove(file_path)

        elif file_name.lower().endswith('.pdf'):
            pdf_document = fitz.open(file_path)
            pdf_text = ""

            for page_number in range(pdf_document.page_count):
                page = pdf_document[page_number]
                pdf_text += page.get_text()

            pdf_document.close()
            print(f'PDF Content:\n{pdf_text}\n{"=" * 50}\n')
            fh.close()
            os.remove(file_path)

        elif file_name.lower().endswith('.py'):
            with open(file_path, 'r', encoding='utf-8') as python_file:
                python_code = python_file.read()

            print(f'Python File Content:\n{python_code}\n{"=" * 50}\n')
            fh.close()
            os.remove(file_path)

    def run_parallel(self, folders):
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(self.content_extract, folders)


if __name__ == "__main__":
    start_time = time.time()

    gdrive_instance = GoogleDriveExtractor()
    folder_list = gdrive_instance.list_folders()
    gdrive_instance.run_parallel(folder_list)

    end_time = time.time()
    print("Total Execution Time:", end_time - start_time)
