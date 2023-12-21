import concurrent
import time
import email
import imaplib
from concurrent import futures
from email.header import decode_header
from email.utils import parsedate_to_datetime
import html2text
import json
import fitz  # PyMuPDF
import io
import openpyxl
from docx import Document


def get_emails(imapUserEmail, imapPassword):
    imap_server = "imap.gmail.com"
    imap_port = 993

    imap_conn = imaplib.IMAP4_SSL(imap_server, imap_port)

    imap_conn.login(imapUserEmail, imapPassword)

    imap_conn.select('inbox')

    _, message_ids = imap_conn.search(None, "UNSEEN")

    emails = []

    for message_id in message_ids[0].split():
        try:
            _, data = imap_conn.fetch(message_id, "(RFC822)")
            raw_email = data[0][1]
            email_message = email.message_from_bytes(raw_email)

            subject = email_message["Subject"]
            sender = email.utils.parseaddr(email_message["From"])[0]
            body1 = ""
            attachments = []

            date_received = parsedate_to_datetime(email_message["Date"])

            if email_message.is_multipart():
                for part in email_message.walk():
                    if part.get_content_type() == "text/plain":
                        body1 = part.get_payload(decode=True).decode()
                    elif part.get_content_type() == "text/html":
                        body1 = html2text.html2text(part.get_payload(decode=True).decode())
                    elif part.get_content_maintype() != 'multipart' and part.get('Content-Disposition') is not None:
                        filename = part.get_filename()
                        if filename:
                            filename, encoding = decode_header(filename)[0]
                            if isinstance(filename, bytes):
                                filename = filename.decode(encoding or "utf-8")
                            attachment_data = part.get_payload(decode=True)
                            attachments.append({
                                "filename": filename,
                                "data": attachment_data
                            })

            email_data = {
                "subject": subject,
                "sender": sender,
                "body": body1,
                "date_received": date_received,
                "email_id": message_id,
                "attachments": attachments
            }

            emails.append(email_data)
        except Exception as e:
            print(f"Error processing email: {str(e)}")

            imap_conn.store(message_id, '-FLAGS', '(\Seen)')

    imap_conn.close()

    return emails


def op_process_mail(email_data):
    subject = email_data["subject"]
    sender = email_data["sender"]
    date_received = email_data["date_received"]
    attachments = email_data["attachments"]
    body = email_data["body"]

    print("Subject:", subject)
    print("Sender:", sender)
    print("Date Received:", date_received)

    if attachments:
        for attachment in attachments:
            filename = attachment["filename"]
            file_format = filename.split('.')[-1].lower()
            content = attachment["data"]

            if file_format == 'xlsx':
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                text_content = ""
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows():
                        row_values = [str(cell.value) for cell in row if cell.value is not None]
                        text_content += " ".join(row_values) + "\n"
                    text_content += "\n"

                print("Attachments:")
                print(json.dumps({
                    "filename": filename,
                    "content_type": "text_from_xlsx",
                    "content": text_content
                }, indent=2))

            elif file_format == 'pdf':
                pdf_document = fitz.open(stream=io.BytesIO(content), filetype="pdf")
                text_content = ""
                for page_number in range(pdf_document.page_count):
                    page = pdf_document[page_number]
                    text_content += page.get_text()

                print("Attachments:")
                print(json.dumps({
                    "filename": filename,
                    "content_type": "text_from_pdf",
                    "content": text_content
                }, indent=2))

            elif file_format == 'docx':
                doc = Document(io.BytesIO(content))
                text_content = "\n".join([paragraph.text for paragraph in doc.paragraphs])

                print("Attachments:")
                print(json.dumps({
                    "filename": filename,
                    "content_type": "text_from_docx",
                    "content": text_content
                }, indent=2))

            elif file_format == 'py':
                text_content = content.decode('utf-8')

                print("Attachments:")
                print(json.dumps({
                    "filename": filename,
                    "content_type": "text_from_python",
                    "content": text_content
                }, indent=2))

    else:
        if body:
            print("Body:", body)
        else:
            print("No Body content")

    print("\n" + "*" * 100 + "\n")


def run_parallel_mails(emails):
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(op_process_mail, emails)


start = time.time()

imapUserEmail = "kiruthika.b@vrdella.com"
imapPassword = "renm kixf nlxy avbx"

emails = get_emails(imapUserEmail, imapPassword)
run_parallel_mails(emails)

end = time.time()
print("Total Execution Time:", end - start)




