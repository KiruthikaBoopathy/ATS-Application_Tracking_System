import concurrent.futures
import time
import email
import imaplib
from concurrent import futures
from email.header import decode_header
from email.utils import parsedate_to_datetime
import html2text
import json
import fitz  # PyMuPDF
import io
import openpyxl
from docx import Document


def get_emails(imapUserEmail, imapPassword):
    imap_server = "imap.gmail.com"
    imap_port = 993

    imap_conn = imaplib.IMAP4_SSL(imap_server, imap_port)

    imap_conn.login(imapUserEmail, imapPassword)

    imap_conn.select('inbox')

    _, message_ids = imap_conn.search(None, "UNSEEN")

    emails = []

    for message_id in message_ids[0].split():
        try:
            _, data = imap_conn.fetch(message_id, "(RFC822)")
            raw_email = data[0][1]
            email_message = email.message_from_bytes(raw_email)

            subject = email_message["Subject"]
            sender = email.utils.parseaddr(email_message["From"])[0]
            body1 = ""
            attachments = []

            date_received = parsedate_to_datetime(email_message["Date"])

            if email_message.is_multipart():
                for part in email_message.walk():
                    if part.get_content_type() == "text/plain":
                        body1 = part.get_payload(decode=True).decode()
                    elif part.get_content_type() == "text/html":
                        body1 = html2text.html2text(part.get_payload(decode=True).decode())
                    elif part.get_content_maintype() != 'multipart' and part.get('Content-Disposition') is not None:

                        filename = part.get_filename()
                        if filename:
                            filename, encoding = decode_header(filename)[0]
                            if isinstance(filename, bytes):
                                filename = filename.decode(encoding or "utf-8")
                            attachment_data = part.get_payload(decode=True)
                            attachments.append({
                                "filename": filename,
                                "data": attachment_data
                            })

            email_data = {
                "subject": subject,
                "sender": sender,
                "body": body1,
                "date_received": date_received,
                "email_id": message_id,
                "attachments": attachments
            }

            emails.append(email_data)
        except Exception as e:
            print(f"Error processing email: {str(e)}")

            imap_conn.store(message_id, '-FLAGS', '(\Seen)')

    imap_conn.close()

    return emails


def extract_text_from_pdf(attachment_data):
    try:
        pdf_document = fitz.open(stream=io.BytesIO(attachment_data), filetype="pdf")

        text_content = ""
        for page_number in range(pdf_document.page_count):
            page = pdf_document[page_number]
            text_content += page.get_text()

        return text_content
    except Exception as e:
        print(f"Error extracting text from PDF: {str(e)}")
        return ""


def handle_xlsx_attachment(attachment_data):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(attachment_data), read_only=True)
        text_content = ""
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                row_values = [str(cell.value) for cell in row if cell.value is not None]
                text_content += " ".join(row_values) + "\n"
            # Add a single newline character after each row
            text_content += "\n"
        return text_content
    except Exception as e:
        print(f"Error handling XLSX attachment: {str(e)}")
        return ""



def handle_docx_attachment(attachment_data):
    try:
        doc = Document(io.BytesIO(attachment_data))
        text_content = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text_content
    except Exception as e:
        print(f"Error handling DOCX attachment: {str(e)}")
        return ""


def handle_python_attachment(attachment_data):
    try:
        return attachment_data.decode('utf-8')
    except Exception as e:
        print(f"Error handling Python script attachment: {str(e)}")
        return ""


def op_process_mail(emails):
    for email_data in emails:
        print("Subject:", email_data["subject"])
        print("Sender:", email_data["sender"])
        print("Date Received:", email_data["date_received"])

        if email_data["attachments"]:
            for attachment in email_data["attachments"]:
                filename = attachment["filename"]
                content = attachment["data"]

                file_format = filename.split('.')[-1].lower()

                if file_format == 'pdf':
                    pdf_text_content = extract_text_from_pdf(content)
                    attachment_info = {
                        "filename": filename,
                        "content_type": "text_from_pdf",
                        "content": pdf_text_content
                    }

                    if attachment_info:
                        print("Attachments:")
                        print(json.dumps(attachment_info, indent=2))
                    else:
                        print("No Attachments")


                elif file_format == 'xlsx':
                    xlsx_text_content = handle_xlsx_attachment(content)
                    attachment_info = {
                        "filename": filename,
                        "content_type": "text_from_xlsx",
                        "content": xlsx_text_content
                    }
                    if attachment_info:
                        print("Attachments:")
                        print(json.dumps(attachment_info, indent=2))
                    else:
                        print("No Attachments")

                elif file_format == 'docx':
                    docx_text_content = handle_docx_attachment(content)
                    attachment_info = {
                        "filename": filename,
                        "content_type": "text_from_docx",
                        "content": docx_text_content
                    }

                    if attachment_info:
                        print("Attachments:")
                        print(json.dumps(attachment_info, indent=2))
                    else:
                        print("No Attachments")

                elif file_format == 'py':
                    python_content = handle_python_attachment(content)
                    attachment_info = {
                        "filename": filename,
                        "content_type": "text_from_python",
                        "content": python_content
                    }

                    if attachment_info:
                        print("Attachments:")
                        print(json.dumps(attachment_info, indent=2))
                    else:
                        print("No Attachments")



        else:
            if email_data["body"]:
                print("Body:", email_data["body"])
            else:
                print("No Body content")

        print("\n")
        print("*" * 100)


def run_parallel(emails):
    futures.ThreadPoolExecutor().map(op_process_mail, emails)
    # with futures.ThreadPoolExecutor() as executor:
    #     executor.map(op_process_mail, emails)


start = time.time()

imapUserEmail = "kiruthika.b@vrdella.com"
imapPassword = "renm kixf nlxy avbx"

emails = get_emails(imapUserEmail, imapPassword)
op_process_mail(emails)
run_parallel(emails)

end = time.time()


print("total:", end - start)
